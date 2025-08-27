import time

from selenium import webdriver
#from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait

import openpyxl


from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook

book: Workbook = openpyxl.load_workbook("C:\\Users\\Khan Gadget\\Desktop\\PythonDemo.xlsx")

# Excel active
sheet: Worksheet = book.active
Dict = {}

for i in range(1, sheet.max_column+1):
    if sheet.cell(row=1, column=i).value == "price":
        Dict["col"] = i

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        if sheet.cell(row=i, column=j).value == "Apple":
            Dict["row"] =i




file_path = "C:\\Users\\Khan Gadget\\Desktop\\download.xlsx"
fruit_name = "Apple"
# chrome driver service
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.maximize_window()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()

# uploaded excel file

sheet.cell(row=Dict["row"], column=Dict["col"]).value = "500"
book.save(file_path)
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)


wait = WebDriverWait(driver, 10)  # explicit wait
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))

print(driver.find_element(*toast_locator).text)

priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")

actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
print(actual_price)