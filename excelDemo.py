import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook

book: Workbook = openpyxl.load_workbook("C:\\Users\\Khan Gadget\\Desktop\\PythonDemo.xlsx")

# Excel active
sheet: Worksheet = book.active
Dict = {}

# kon column r row niye kaj hobe seita bola hoise seita niye
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = "shadman"

print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)

print(sheet.max_column)

print(sheet['A5'].value)

for i in range(1, sheet.max_row+1):  # to get rows
    if sheet.cell(row=i, column=1).value == "Testcase2":

        for j in range(1, sheet.max_column+1):

            print(sheet.cell(row=i, column=j).value)